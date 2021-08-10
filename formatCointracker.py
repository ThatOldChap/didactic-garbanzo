#! python3
# formatCointracker.py
# Formats various exchanges crypto transactions into the Cointracker format

import os, openpyxl, shutil, csv
from pathlib import Path
from openpyxl import workbook 
from datetime import datetime

# Supported Exchanges
COINSQUARE = 0
NDAX = 1

# Supported Report Types
FUND_AND_WITHDRAW = 2
QUICK_TRADE = 3

# Exchange Fees
COINSQUARE_BTC_TX_FEE = .002            # 0.2%
COINSQUARE_NON_BTC_TX_FEE = .004        # 0.4%
COINSQUARE_BTC_WITHDRAW_FEE = 0.0005
COINSQUARE_ETH_WITHDRAW_FEE = 0.005
COINSQUARE_DOGE_WITHDRAW_FEE = 2

# Assorted
BUY_TX = 4
SELL_TX = 5

# Lists of letters
A_TO_G_LIST = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
A_TO_H_LIST = A_TO_G_LIST + ['H']
I_TO_L_LIST = ['I', 'J', 'K', 'L']
A_TO_G_NO_I_LIST = A_TO_G_LIST + I_TO_L_LIST

# Convert .csv files to .xlsx
def csvToXlsx(csv_file_path):

    # Create the excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"

    # Read/Write the file
    with open(csv_file_path, 'rt', encoding='utf8') as file:
        reader = csv.reader(file)
        for row, row_val in enumerate(reader):
            # Offsets the rows since Excel starts at A=1
            adj_row = row + 1
            for col, col_val in enumerate(row_val):
                # Offsets the columns since Excel starts at A=1
                adj_col = col + 1
                worksheet.cell(row=adj_row, column=adj_col).value = col_val
    
    # Organize the files
    xlsx_file_dir = os.path.abspath(csv_file_path.parent) + '\\' + csv_file_path.stem + '.xlsx'
    xlsx_file_path = Path(xlsx_file_dir)
    workbook.save(os.path.abspath(xlsx_file_path))
    os.unlink(csv_file_path)
    print(f'Successfully converted {csv_file_path.name} to {xlsx_file_path.name}')


# Returns the exchange from which the report file originated from
def getExchangeName(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Determine the exchange based on the report file format
    if sheet['A1'].value == 'txid':
        return NDAX
    elif sheet['G1'].value == 'btid' or sheet['E1'].value == 'to_amount':
        return COINSQUARE


# Returns the type of report generated from Coinsquare
def getCoinsquareReportType(sheet):
    if sheet['B1'].value == "description":
        return FUND_AND_WITHDRAW
    elif sheet['B1'].value == "from_currency":
        return QUICK_TRADE


# Get a list of file paths given a set of search criteria
def getFilePathListDict(file_path, file_ext_list):
    file_path_list_dict = {}
    for ext in file_ext_list:
        file_path_list_dict[ext] = list(file_path.glob('*.' + ext))
    return file_path_list_dict        


# Writes data to a spreadsheet using lists of indices
def writeToExcelSheet(sheet, col_letters, row_number, data):
    for i in range(len(data)):
        sheet[col_letters[i] + str(row_number)] = data[i]


# Formats the header row of the Transcation files
def formatCointrackerHeader(sheet):
    col_names = ['Date', 'Received Quantity', 'Received Currency', 'Sent Quantity',
        'Sent Currency', 'Fee Amount', 'Fee Currency', 'Tag']
    writeToExcelSheet(sheet, A_TO_H_LIST, 1, col_names)


# Create a directory for a supplied folder path
def createDirectory(folder_path):
    if not folder_path.exists():
        os.makedirs(folder_path)
        print(f'Successfully generated directory at {folder_path}')


# Create a Master Ledger file
def createMasterLedger(file_path):
    if not file_path.is_file():
        # Create the workbook and worksheet
        ledger_workbook = openpyxl.Workbook()
        ledger_sheet = ledger_workbook.active
        ledger_sheet.title = "Transactions"
        ledger_sheet = ledger_workbook["Transactions"]    

        # Worksheet formatting    
        formatCointrackerHeader(ledger_sheet)
        extra_col_names = ['Cost Basis', 'Cost Basis Units', 'Exchange', 'Tx_Id']
        writeToExcelSheet(ledger_sheet, I_TO_L_LIST, 1, extra_col_names)

        # Save the Master Ledger file
        ledger_workbook.save(os.path.abspath(file_path))
        print(f'Successfully created new Master Ledger file at {file_path}')


def formatCoinsquareDate(date):
    date_split = date.split("-")
    day = date_split[0]
    month = date_split[1]
    year = date_split[2]
    return month + '/' + day + '/20' + year + ' 13:00:00'


def formatNDAXDate(raw_date, raw_time):
    # Date = 2021-07-28
    date_split = raw_date.split("-")
    year = date_split[0]
    month = date_split[1]
    day = date_split[2]    

    # Time = 5:31 PM
    raw_hour = raw_time.split(":")
    raw_other = raw_hour[1].split(" ")
    AM_or_PM = raw_other[1]
    minute = raw_other[0]
    hour = raw_hour[0]
    hour_int = int(hour)

    # Convert from 12 hour to 24 hour time format
    if AM_or_PM == "PM" and hour_int < 12:
        hour = hour_int + 12

    return month + '/' + day + '/' + year + ' ' + str(hour) + ':' + minute + ':' + '00'


# Calculates the Coinsquare transaction fees that aren't explcitly in the reports
def calcCoinsquareFee(received_qty, received_currency, sent_qty, sent_currency):

    # Determines the transaction fees based on the currency used
    if received_currency == "BTC" or sent_currency == "BTC":
        received_fee_amount = received_qty * COINSQUARE_BTC_TX_FEE        
        sent_fee_amount = sent_qty * COINSQUARE_BTC_TX_FEE
    else:
        received_fee_amount = received_qty * COINSQUARE_NON_BTC_TX_FEE        
        sent_fee_amount = sent_qty * COINSQUARE_NON_BTC_TX_FEE
    
    # Makes the transaction fee available as both in sent and received currencies
    received_fee_currency = received_currency
    sent_fee_currency = sent_currency

    return {"received_fee_amount": received_fee_amount, "received_fee_currency": received_fee_currency,
            "sent_fee_amount": sent_fee_amount, "sent_fee_currency": sent_fee_currency}


# Calculates the cost basis of a transaction
def calcTxCostBasis(received_qty, received_currency, sent_qty, sent_currency, fee_amount, tx_type):
    if tx_type == BUY_TX:
        # Since the received_qty is before the fee_amount is subtracted, the fee has to be subtracted now
        return {"cost_basis": round(sent_qty/(received_qty - fee_amount), 3),
                "cost_basis_units": sent_currency + "/" + received_currency}
    elif tx_type == SELL_TX:
        return {"cost_basis": round(received_qty/sent_qty, 3),
                "cost_basis_units": received_currency + "/" + sent_currency}


# Determines the type of transaction based on trading between CAD and Crypto
# TODO: Future update for once I trade between cryptos
def getTradeType(received_currency, sent_currency):
    if received_currency == "CAD":
        return SELL_TX
    elif sent_currency == "CAD":
        return BUY_TX


# Extracts the number from a spreadsheet cell that is stored as text (ie. removes commas for pure decimal)
def extractFloatFromText(text):
    return float(text.replace(',',''))


# Generates a new timestamped, unique filename
def generateFilename(name, file_ext):
    timestamp_obj = datetime.now()
    timestamp = timestamp_obj.strftime("%d%b%Y_%H%M%S%f")
    filename = name + "_Txs_" + timestamp + file_ext
    return filename


# Saves a newly formatted results file
def saveNewResultFile(report_workbook, report_path, new_file_path, filename):    
    # Save the workbook at the provided path
    report_workbook.save(os.path.abspath(report_path))    

    # Move the new file to the new filepath and rename
    shutil.move(os.path.abspath(report_path), os.path.abspath(new_file_path))
    print(f'Saved and moved new file as {filename}.')


# Add a transaction to the Master Ledger data list
def addMasterLedgerData(data, date, exchange, received_qty, received_currency,
                    sent_qty, sent_currency, fee_amount, fee_currency, cost_basis,
                    cost_basis_units, tx_id):
    data.append({
        "date": date,
        "exchange": exchange,
        "received_qty": received_qty,
        "received_currency": received_currency,
        "sent_currency": sent_currency,
        "sent_qty": sent_qty,
        "fee_amount": fee_amount,
        "fee_currency": fee_currency,
        "cost_basis": cost_basis,
        "cost_basis_units": cost_basis_units,
        "tx_id": tx_id
    })


# Update the Master Ledger with a new dataset
def updateMasterLedger(data, ledger_path):

    # Open the existing Master Ledger file and worksheet
    ledger_workbook = openpyxl.load_workbook(ledger_path)
    ledger_sheet = ledger_workbook.active

    # Find the next row available to start adding transactions after
    last_tx_row = ledger_sheet.max_row
    num_tx = len(data)
    row_start = last_tx_row + 1
    row_end = row_start + num_tx
    print(f'Starting update of Master Ledger on row {row_start} with {num_tx} transactions.')

    # TODO: Sort the transactions by date
    # TODO: Lookup previous transactions in the master ledger first by date, then loop through to find exact matches
    tx_check = 0
    dupe_tx = 0

    # Check the existing Master Ledger for the current transaction ids (tx_id)
    tx_id_list = getAllTxIDs(ledger_path)
    tx_id_exists = False

    for i in range(row_start, row_end):
        # Adjust xl_offset to continue writing to the last_row of the Master Ledger avoiding spaces from dupe_tx
        xl_offset = i - dupe_tx

        # Adjust data_offset to start at index 0 of the data array
        data_offset = xl_offset - row_start
        
        # Check if the transaction id already exists in the data collected
        tx_id_exists = tx_id_list.count(data[data_offset]["tx_id"]) > 0

        if tx_id_exists:
            # tx_id exists so skip re-writing it to avoid duplication
            dupe_tx += 1

            # Remove the duplicate data from the ledger data
            # Note: Same data[index] needs to be checked again after dupe_tx popped from data
            data.pop(data_offset)
            continue

        elif not tx_id_exists:
            # Write the data to the ledger's worksheet
            ledger_data = [data[data_offset]["date"],   data[data_offset]["received_qty"],
                data[data_offset]["received_currency"], data[data_offset]["sent_qty"],
                data[data_offset]["sent_currency"],     data[data_offset]["fee_amount"],
                data[data_offset]["fee_currency"],      data[data_offset]["cost_basis"],
                data[data_offset]["cost_basis_units"],  data[data_offset]["exchange"],
                data[data_offset]["tx_id"]]
            # Excludes writing to the 'Tag' column since there aren't any I've used yet
            writeToExcelSheet(ledger_sheet, A_TO_G_NO_I_LIST, xl_offset, ledger_data)
            tx_check += 1            
    
    # Save the updated Master Ledger
    ledger_workbook.save(os.path.abspath(ledger_path))
    if tx_check + dupe_tx - num_tx == 0:
        print(f'Successfully updated the Master Ledger with {tx_check} new transactions out of {num_tx} total.')
    else:
        print(f'Error updating the Master Ledger. Total of {tx_check+dupe_tx}/{num_tx} transactions completed.')


# Format a Coinsquare report of type: FUND_AND_WITHDRAW
def formatFundAndWithdrawReport(data, raw_sheet, new_sheet, last_row):
    # TODO: Add in a future custom Coinsquare tx_id
    tx_id = ""
    exchange = "Coinsquare"

    # Read and format the data from the file
    for i in range(2, last_row+1):
        # Date
        date = formatCoinsquareDate(raw_sheet['A' + str(i)].value)

        # Amount Info             
        qty = extractFloatFromText(raw_sheet['D' + str(i)].value)   # Amount           
        currency = raw_sheet['E' + str(i)].value                    # Currency
        operation = raw_sheet['C' + str(i)].value                   # Credit or Debit
        cost_basis = ""
        cost_basis_units = ""

        # Calculate fees and quantities
        if operation == "credit":
            received_qty = qty
            received_currency = currency
            sent_qty = ""
            sent_currency = ""
            fee_amount = ""
            fee_currency = ""
        elif operation == "debit":
            if currency == "BTC":
                fee_amount = COINSQUARE_BTC_WITHDRAW_FEE
            if currency == "ETH":
                fee_amount = COINSQUARE_ETH_WITHDRAW_FEE
            if currency == "DOGE":
                fee_amount = COINSQUARE_DOGE_WITHDRAW_FEE   
            sent_qty = qty              # Includes fees 
            sent_currency = currency                 
            fee_currency = currency  

            # Reformats a transfer as a send/receive of the same amount, with a fee deducted        
            received_qty = sent_qty
            received_currency = sent_currency

        # Write the data to the new_sheet on the original file
        new_data = [date, received_qty, received_currency, sent_qty, sent_currency, fee_amount, fee_currency]
        writeToExcelSheet(new_sheet, A_TO_G_LIST, i, new_data)

        # Add the new data to be list of possible new transactions for the master ledger
        addMasterLedgerData(data, date, exchange, received_qty, received_currency,
            sent_qty, sent_currency, fee_amount, fee_currency, cost_basis,
            cost_basis_units, tx_id)


# Format a Coinsquare report of type: QUICK_TRADE
def formatQuickTradeReport(data, raw_sheet, new_sheet, last_row):
    # TODO: Add in a future custom Coinsquare tx_id
    tx_id = ""
    exchange = "Coinsquare"

    for i in range(2, last_row+1):
        # Date
        date = formatCoinsquareDate(raw_sheet['A' + str(i)].value)

        # From/To Info
        from_currency = raw_sheet['B' + str(i)].value
        from_amount = extractFloatFromText(raw_sheet['C' + str(i)].value)
        to_currency = raw_sheet['D' + str(i)].value
        to_amount = extractFloatFromText(raw_sheet['E' + str(i)].value)

        fee_dict = calcCoinsquareFee(to_amount, to_currency, from_amount, from_currency)

        # Calculate the cost_basis and determine the fee_currency
        tx_type = getTradeType(to_currency, from_currency)

        if tx_type == BUY_TX:
            fee_currency = fee_dict['received_fee_currency']
            fee_amount = fee_dict['received_fee_amount']

            # Adjust the received_qty to include the fee
            to_amount += fee_amount

            cost_basis_dict = calcTxCostBasis(to_amount, to_currency, from_amount, from_currency, fee_amount, BUY_TX)

        elif tx_type == SELL_TX:
            fee_currency = fee_dict['sent_fee_currency']
            fee_amount = fee_dict['sent_fee_amount']

            cost_basis_dict = calcTxCostBasis(to_amount, to_currency, from_amount, from_currency, fee_amount, SELL_TX)
        else:
            print('Error calculating fees and cost basis')

        # Extract the cost basis info
        cost_basis = cost_basis_dict["cost_basis"]
        cost_basis_units = cost_basis_dict["cost_basis_units"]

        # Write the data to the new_sheet on the original file
        row_data = [date, to_amount, to_currency, from_amount, from_currency, fee_amount, fee_currency]     
        writeToExcelSheet(new_sheet, A_TO_G_LIST, i, row_data)

        # Add the new data to be list of possible new transactions for the master ledger
        addMasterLedgerData(data, date, exchange, to_amount, to_currency,
            from_amount, from_currency, fee_amount, fee_currency,
            cost_basis, cost_basis_units, tx_id)


# Format a Coinsquare Ledger file for the "Fund/Withdraw" transactions
def formatCoinsquare(data, report_path, new_file_dir):
    print(f'Preparing to format file at {report_path}...')

    # Generate a filename
    exchange = "Coinsquare"
    filename = generateFilename(exchange, '.xlsx')
    new_file_path = Path(new_file_dir + filename)

    # Load the file being processed    
    report_workbook = openpyxl.load_workbook(report_path)

    # Defines the sheets
    raw_sheet = report_workbook.active
    new_sheet = report_workbook.create_sheet(title='Formatted')
    last_row = raw_sheet.max_row
    formatCointrackerHeader(new_sheet)

    # Determine the report type
    report_type = getCoinsquareReportType(raw_sheet)
    if report_type == FUND_AND_WITHDRAW:
        formatFundAndWithdrawReport(data, raw_sheet, new_sheet, last_row)
    elif report_type == QUICK_TRADE:
        formatQuickTradeReport(data, raw_sheet, new_sheet, last_row)

    # Save the new formatted file
    saveNewResultFile(report_workbook, report_path, new_file_path, filename)
    return data


# Format an NDAX transactions file
def formatNDAX(data, report_path, new_file_dir):    
    print(f'Preparing to format file at {report_path}...')

    # Generate a filename
    exchange = "NDAX"
    filename = generateFilename(exchange, '.xlsx')
    new_file_path = Path(new_file_dir + filename)

    # Load the file being processed    
    report_workbook = openpyxl.load_workbook(report_path)

    # Defines the sheets
    raw_sheet = report_workbook.active
    new_sheet = report_workbook.create_sheet(title='Formatted')
    last_row = raw_sheet.max_row
    formatCointrackerHeader(new_sheet)

    # Define the tracking index to help process the multi-line nature of the transactions
    next_tx_index = last_row
    skip_offset = 0

    # Read and format the data from the file
    for i in range(last_row, 2, -1):

        # Skip indices to get to the next transaction after a trade has been processed
        if i > next_tx_index:
            continue

        # Offset for starting to write data at row 2 in the new formatted sheet
        offset = last_row - i + 2 - skip_offset

        # Reference IDs
        tx_id = raw_sheet['A' + str(i)].value
        # ref_id = raw_sheet['B' + str(i)].value

        # Datetime
        raw_date = raw_sheet['C' + str(i)].value
        raw_time = raw_sheet['D' + str(i)].value
        date = formatNDAXDate(raw_date, raw_time)

        # Type and Cost Basis
        tx_type = raw_sheet['E' + str(i)].value
        cost_basis = ""
        cost_basis_units = ""

        # Finds the relevant transactions and fees based on the type of transaction
        if tx_type == 'Deposit':
            # Get the deposit info
            received_qty = float(raw_sheet['H' + str(i)].value)
            received_currency = raw_sheet['F' + str(i)].value

            # Process fiat currency (CAD) deposits only
            if received_currency == "CAD":
                sent_qty = ""
                sent_currency = ""
                fee_amount = ""
                fee_currency = ""

            # Ignore non-fiat transfers because transfers are dealt with as tx_type=Trade
            else:                
                # Assign the indices for the current transaction to skip over the next loop
                skip_offset += 1
                continue

        elif tx_type == 'Affiliate Payout':
            received_qty = float(raw_sheet['H' + str(i)].value)
            received_currency = raw_sheet['F' + str(i)].value
            sent_qty = ""
            sent_currency = ""
            fee_amount = ""
            fee_currency = ""
            
        elif tx_type == 'Trade':
            # Fees
            qty_i_2 = raw_sheet['H' + str(i-2)].value
            fee_amount = float(qty_i_2)*-1
            fee_currency = raw_sheet['F' + str(i-2)].value

            # Determine which indexes are the sent/receive
            qty_i = raw_sheet['H' + str(i)].value
            qty_i_1 = raw_sheet['H' + str(i-1)].value
            if float(qty_i) > 0:
                received_qty = float(qty_i)
                received_currency = raw_sheet['F' + str(i)].value
                sent_qty = float(qty_i_1)*-1
                sent_currency = raw_sheet['F' + str(i-1)].value
            elif float(qty_i) < 0:
                received_qty = float(qty_i_1)
                received_currency = raw_sheet['F' + str(i-1)].value
                sent_qty = float(qty_i)*-1
                sent_currency = raw_sheet['F' + str(i)].value

            # Assign the indices for the current transaction to skip over the next 2 loops
            next_tx_index = i - 3
            skip_offset += 2
        
            # Determine the cost basis for the transaction
            trade_type = getTradeType(received_currency, sent_currency)
            cost_basis_dict = calcTxCostBasis(received_qty, received_currency, sent_qty, sent_currency,
                                                fee_amount, trade_type)
            cost_basis = cost_basis_dict["cost_basis"]
            cost_basis_units = cost_basis_dict["cost_basis_units"]        

        # Write the data to the new_sheet on the original file
        new_data = [date, received_qty, received_currency, sent_qty, sent_currency, fee_amount, fee_currency]
        writeToExcelSheet(new_sheet, A_TO_G_LIST, offset, new_data)

        # Add the new data to be list of possible new transactions for the master ledger
        addMasterLedgerData(data, date, exchange, received_qty, received_currency,
            sent_qty, sent_currency, fee_amount, fee_currency, cost_basis,
            cost_basis_units, tx_id)
    
    saveNewResultFile(report_workbook, report_path, new_file_path, filename)
    return data


# Returns a list of all the transaction ids (tx_id) present in the Master Ledger
def getAllTxIDs(ledger_path):
    # Load the Master Ledger worksheet
    workbook = openpyxl.load_workbook(ledger_path)
    sheet = workbook.active

    # Create a list of all the tx_ids present
    tx_id_list = []
    last_tx_row = sheet.max_row
    row_start = 2   # Start on row 2 to bypass header row
    row_end = last_tx_row + 1
    
    for i in range(row_start, row_end):
        tx_id = sheet['J' + str(i)].value
        if tx_id != "":
            tx_id_list.append(tx_id)

    return tx_id_list


# Creates a summarized import file for Cointracker
def getCointrackerSummary(data, new_file_dir):
    print(f'Preparing Cointracker Summary file at {results_dir}...')

    # Generate a filename
    name = "Cointracker_Import"
    filename = generateFilename(name, '.csv')
    new_file_path = Path(new_file_dir + filename)

    # Setup a checker to confirm all transactions were processed
    tx_check = 0
    num_tx = len(data)

    # Create a new .csv file
    with open(new_file_path, 'w', newline='') as file:
        writer = csv.writer(file)

        # Write header row
        writer.writerow(["Date", "Received Quantity", "Received Currency", "Sent Quantity",
            "Sent Currency", "Fee Amount", "Fee Currency", "Tag"])
        
        # Write all the data to the file
        for i in range(len(data)):
            row = data[i]
            writer.writerow([
                row["date"],
                row["received_qty"],
                row["received_currency"],
                row["sent_qty"],
                row["sent_currency"],
                row["fee_amount"],
                row["fee_currency"]
            ])
            tx_check += 1
    
    if tx_check - num_tx == 0:
        print(f'Successfully created new Cointracker Summary file with {tx_check} transactions.')
    else:
        print(f'Error creating new Cointracker Summary file. Total transactions completed are {tx_check}/{num_tx}.')


# Converts any .csv report files into .xlsx files
def convertCSVFiles(reports_path):
    # Get a dictionary of all the .csv file paths to prep for conversion to .xlsx
    csv_file_paths = getFilePathListDict(reports_path, ['csv'])
    csv_file_paths_list = csv_file_paths["csv"]

    if len(csv_file_paths_list) == 0:
        return
    else:
        # Convert all .csv files into .xlsx
        for csv_file_path in csv_file_paths_list:
            csvToXlsx(csv_file_path)


# Main method for processing all the exchange's reports
def processReports(reports_path, results_dir, ledger_path):

    # Convert any .csv reports if applicable
    convertCSVFiles(reports_path)

    # Get a dictionary of all the .xlsx file paths to prep for analysis
    xlsx_file_paths = getFilePathListDict(reports_path, ['xlsx'])
    xlsx_file_paths_list = xlsx_file_paths["xlsx"]

    # Start a new dataset for all the transactions of the newly processed files
    data = []

    for report_path in xlsx_file_paths_list:
        exchange = getExchangeName(report_path)
        if exchange == COINSQUARE:
            formatCoinsquare(data, report_path, results_dir)
        elif exchange == NDAX:
            formatNDAX(data, report_path, results_dir)

    # Update the Master Ledger with the new data
    updateMasterLedger(data, ledger_path)

    # Create a summarized import form for all new transactions to import into Cointracker
    getCointrackerSummary(data, results_dir)


# Initialize the directory structure and create a Master Ledger
def init(reports_path, results_path, ledger_path):
    # Setup the directories for the exchange's reports and the formatted results
    createDirectory(reports_path)
    createDirectory(results_path)

    # Create a Master Ledger file
    createMasterLedger(ledger_path)


# Main crypto directory, which will also store the Master Ledger
crypto_dir = 'C:\\Users\\michael chaplin\\OneDrive - MDS Aero Support\\Documents\\Python\\Crypto\\'
ledger_dir = crypto_dir + 'Master_Ledger.xlsx'
ledger_path = Path(ledger_dir)

# Directory to store the exchange's reports 
reports_dir = crypto_dir + 'Reports\\'
reports_path = Path(reports_dir)

# Directory to store the formatted result files
results_dir = crypto_dir + 'Results\\'
results_path = Path(results_dir)

# Initialize the directory and Master Ledger
init(reports_path, results_path, ledger_path)

# Process each of the exchange's reports
processReports(reports_path, results_dir, ledger_path)
